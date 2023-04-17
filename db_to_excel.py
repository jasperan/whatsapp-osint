from openpyxl import Workbook
import sqlite3
from openpyxl.styles import Font, Alignment


class Converter:
    def db_to(self):
        self.conn = sqlite3.connect('database/database.db')
        self.cursor = self.conn.cursor()
        self.wb = Workbook()
        self.ws = self.wb.active

    def db_to_excel(self):
        # Style
        bold = Font(bold=True, name='Arial', color="00800000", size=10)
        align = Alignment(horizontal="center")

        # Default Column
        self.ws.column_dimensions['A'].width = 15
        self.ws["A1"].font = bold
        self.ws["A1"].alignment = align
        self.ws["A1"] = "Id"

        self.ws.column_dimensions['B'].width = 17
        self.ws["B1"].font = bold
        self.ws["B1"].alignment = align
        self.ws["B1"] = "Username"

        self.ws.column_dimensions['C'].width = 15
        self.ws["C1"].font = bold
        self.ws["C1"].alignment = align
        self.ws["C1"] = "DateTime"

        self.ws.column_dimensions['D'].width = 15
        self.ws["D1"].font = bold
        self.ws["D1"].alignment = align
        self.ws["D1"] = "Hour"

        self.ws.column_dimensions['E'].width = 15
        self.ws["E1"].font = bold
        self.ws["E1"].alignment = align
        self.ws["E1"] = "Minutes"

        self.ws.column_dimensions['F'].width = 15
        self.ws["F1"].font = bold
        self.ws["F1"].alignment = align
        self.ws["F1"] = "Second"

        self.ws.column_dimensions['G'].width = 17
        self.ws["G1"].font = bold
        self.ws["G1"].alignment = align
        self.ws["G1"] = "Type_Connection"

        self.ws.column_dimensions['H'].width = 15
        self.ws["H1"].font = bold
        self.ws["H1"].alignment = align
        self.ws["H1"] = "Online Time"

        self.ws.title = "History Of Their Wp  "

        try:
            self.cursor.execute("Select * From Connections Order by date DESC ")
            all_data = self.cursor.fetchall()

            for data in all_data:
                self.ws.append(data)
            print("All data added your excel file")
            self.wb.save("History_wp.xlsx")

        except PermissionError:
            print("Please Close Your Excel File and restart  ")


