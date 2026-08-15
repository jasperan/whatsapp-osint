from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sqlite3
import logging
import json
from datetime import datetime
from pathlib import Path

from .database import Database

logger = logging.getLogger(__name__)

class Converter:
    def __init__(self, db_path: str = 'data/victims_logs.db', excel_file: str = 'History_wp.xlsx'):
        self.db_path = Path(db_path)
        self.excel_file = Path(excel_file)

    def db_to_excel(self):
        """Exports data from the database to an Excel file."""
        if not self.db_path.exists():
            logger.error(f"Database not found at {self.db_path}")
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                wb = Workbook()
                ws = wb.active

                bold = Font(bold=True, name='Arial', color="00800000", size=10)
                align = Alignment(horizontal="center")

                headers = [
                    ("A", 15, "Session ID"),
                    ("B", 17, "Username"),
                    ("C", 20, "Start DateTime"),
                    ("D", 20, "End DateTime"),
                    ("E", 15, "Time Connected (s)")
                ]

                for col, width, title in headers:
                    ws.column_dimensions[col].width = width
                    cell = ws[f"{col}1"]
                    cell.font = bold
                    cell.alignment = align
                    cell.value = title

                ws.title = "History Of Their Wp"

                query = '''
                    SELECT
                        s.id,
                        u.user_name,
                        s.start_date || ' ' || s.start_hour || ':' || s.start_minute || ':' || s.start_second AS start_datetime,
                        s.end_date || ' ' || s.end_hour || ':' || s.end_minute || ':' || s.end_second AS end_datetime,
                        s.time_connected
                    FROM Sessions s
                    JOIN Users u ON s.user_id = u.id
                    WHERE s.end_date IS NOT NULL
                    ORDER BY s.start_date DESC, s.start_hour DESC, s.start_minute DESC, s.start_second DESC
                '''
                cursor.execute(query)
                all_data = cursor.fetchall()

                # Presence worksheet: passive "last seen" / typing / online evidence.
                presence_headers = [
                    ("A", 17, "Username"),
                    ("B", 20, "Observed At"),
                    ("C", 15, "Status Kind"),
                    ("D", 34, "Status Text"),
                    ("E", 20, "Last Seen (parsed)"),
                ]
                ws2 = wb.create_sheet(title="Presence")
                for col, width, title in presence_headers:
                    ws2.column_dimensions[col].width = width
                    cell = ws2[f"{col}1"]
                    cell.font = bold
                    cell.alignment = align
                    cell.value = title

                presence_data = list(reversed(Database(str(self.db_path)).get_presence_history()))

            for row_idx, data in enumerate(all_data, start=2):
                ws[f"A{row_idx}"] = data[0]
                ws[f"B{row_idx}"] = data[1]
                ws[f"C{row_idx}"] = data[2]
                ws[f"D{row_idx}"] = data[3]
                ws[f"E{row_idx}"] = data[4]

            for row_idx, data in enumerate(presence_data, start=2):
                ws2[f"A{row_idx}"] = data['user_name']
                ws2[f"B{row_idx}"] = data['observed_at']
                ws2[f"C{row_idx}"] = data['status_kind']
                ws2[f"D{row_idx}"] = data['status_text']
                ws2[f"E{row_idx}"] = data['last_seen']

            try:
                self.excel_file.parent.mkdir(parents=True, exist_ok=True)
                wb.save(self.excel_file)
                logger.info(f"All data added to your Excel file: {self.excel_file}")
            except PermissionError:
                logger.error(f"Please close '{self.excel_file}' and restart the program.")

        except sqlite3.Error as e:
            logger.error(f"Database error: {e}")


def export_json(db_path: str = 'data/victims_logs.db', output_path: str = 'analytics/sessions.json') -> Path:
    """Writes sessions and presence signals to a JSON file for pipelines.

    Returns the written path.
    """
    out = Path(output_path)
    db = Path(db_path)
    payload = {
        'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'sessions': [],
        'presence': [],
    }
    if db.exists():
        try:
            with sqlite3.connect(str(db)) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT
                        u.user_name,
                        s.start_date || ' ' || s.start_hour || ':' || s.start_minute || ':' || s.start_second AS start_datetime,
                        s.end_date || ' ' || s.end_hour || ':' || s.end_minute || ':' || s.end_second AS end_datetime,
                        s.time_connected
                    FROM Sessions s
                    JOIN Users u ON s.user_id = u.id
                    ORDER BY s.start_date ASC, s.start_hour ASC, s.start_minute ASC, s.start_second ASC
                ''')
                for row in cursor.fetchall():
                    payload['sessions'].append({
                        'user_name': row[0],
                        'start_datetime': row[1],
                        'end_datetime': row[2],
                        'time_connected': row[3],
                    })
        except sqlite3.Error as e:
            logger.error(f"Database error during JSON export: {e}")
        # Presence rows come from the canonical Database accessor so the query
        # and row shape stay defined in exactly one place.
        payload['presence'] = Database(db_path=str(db)).get_presence_history()

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    logger.info(f"JSON export written to {out}")
    return out
