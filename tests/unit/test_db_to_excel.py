"""Unit tests for Excel + JSON exports (incl. the Presence worksheet)."""
import json

import pytest
from openpyxl import load_workbook

from src.whatsapp_beacon.database import Database
from src.whatsapp_beacon.db_to_excel import Converter, export_json


def _seed_db(db_path, tmp_path):
    db = Database(db_path=str(db_path))
    alice = db.get_or_create_user('Alice')
    session_id = db.insert_session_start(alice, {
        'date': '2026-08-01', 'hour': '10', 'minute': '00', 'second': '00'
    })
    db.update_session_end(session_id, {
        'date': '2026-08-01', 'hour': '10', 'minute': '05', 'second': '00'
    }, '300')
    db.insert_presence(
        user_id=alice,
        observed_at='2026-08-02 09:00:00',
        status_kind='last_seen',
        status_text='last seen yesterday at 22:00',
        last_seen='2026-08-01 22:00:00',
    )
    return db


def test_db_to_excel_writes_sessions_and_presence_sheets(tmp_path):
    db_path = tmp_path / 'export.db'
    _seed_db(db_path, tmp_path)
    excel_path = tmp_path / 'out' / 'History_wp.xlsx'

    Converter(db_path=str(db_path), excel_file=str(excel_path)).db_to_excel()

    assert excel_path.exists()
    wb = load_workbook(excel_path)
    assert wb.sheetnames == ['History Of Their Wp', 'Presence']
    sessions_sheet = wb['History Of Their Wp']
    assert sessions_sheet['A1'].value == 'Session ID'
    assert sessions_sheet['B2'].value == 'Alice'
    assert sessions_sheet['E2'].value == '300'

    presence_sheet = wb['Presence']
    assert presence_sheet['A1'].value == 'Username'
    assert presence_sheet['B2'].value == '2026-08-02 09:00:00'
    assert presence_sheet['D2'].value == 'last seen yesterday at 22:00'
    assert presence_sheet['E2'].value == '2026-08-01 22:00:00'


def test_export_json_writes_sessions_and_presence(tmp_path):
    db_path = tmp_path / 'export.db'
    _seed_db(db_path, tmp_path)
    out_path = tmp_path / 'out' / 'sessions.json'

    written = export_json(db_path=str(db_path), output_path=str(out_path))
    assert written == out_path

    payload = json.loads(out_path.read_text(encoding='utf-8'))
    assert set(payload) == {'exported_at', 'sessions', 'presence'}
    assert len(payload['sessions']) == 1
    assert payload['sessions'][0]['user_name'] == 'Alice'
    assert payload['sessions'][0]['time_connected'] == '300'
    assert len(payload['presence']) == 1
    assert payload['presence'][0]['status_kind'] == 'last_seen'


def test_export_json_handles_missing_database(tmp_path):
    out_path = tmp_path / 'missing.json'
    written = export_json(db_path=str(tmp_path / 'nope.db'), output_path=str(out_path))
    payload = json.loads(written.read_text(encoding='utf-8'))
    assert payload['sessions'] == []
    assert payload['presence'] == []


def test_db_to_excel_handles_missing_database(tmp_path):
    converter = Converter(db_path=str(tmp_path / 'nope.db'), excel_file=str(tmp_path / 'out.xlsx'))
    converter.db_to_excel()  # should log and return without raising
    assert not (tmp_path / 'out.xlsx').exists()


def test_db_to_excel_works_with_legacy_db_without_presence_table(tmp_path):
    """A DB created before the Presence feature must still export sessions."""
    import sqlite3

    db_path = tmp_path / 'legacy.db'
    with sqlite3.connect(str(db_path)) as conn:
        conn.execute('''
            CREATE TABLE Users (id INTEGER PRIMARY KEY AUTOINCREMENT, user_name TEXT UNIQUE)
        ''')
        conn.execute('''
            CREATE TABLE Sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
                start_date TEXT, start_hour TEXT, start_minute TEXT, start_second TEXT,
                end_date TEXT, end_hour TEXT, end_minute TEXT, end_second TEXT,
                time_connected TEXT
            )
        ''')
        conn.execute("INSERT INTO Users (user_name) VALUES ('Alice')")
        conn.execute('''
            INSERT INTO Sessions (user_id, start_date, start_hour, start_minute, start_second,
                end_date, end_hour, end_minute, end_second, time_connected)
            VALUES (1, '2026-08-01', '10', '00', '00', '2026-08-01', '10', '05', '00', '300')
        ''')

    excel_path = tmp_path / 'legacy.xlsx'
    Converter(db_path=str(db_path), excel_file=str(excel_path)).db_to_excel()

    assert excel_path.exists()
    wb = load_workbook(excel_path)
    assert wb.sheetnames == ['History Of Their Wp', 'Presence']
    assert wb['History Of Their Wp']['B2'].value == 'Alice'
    assert wb['Presence']['A2'].value is None  # empty presence sheet
