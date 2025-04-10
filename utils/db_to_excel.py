from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sqlite3

class Converter:
    DB_PATH = 'database/victims_logs.db'
    EXCEL_FILE = 'History_wp.xlsx'

    def db_to(self):
        """Inicializa la conexión a la base de datos y crea el libro de Excel."""
        self.conn = sqlite3.connect(self.DB_PATH)
        self.cursor = self.conn.cursor()
        self.wb = Workbook()
        self.ws = self.wb.active

    def db_to_excel(self):
        """Exporta los datos de la base de datos a un archivo Excel."""
        # Estilos
        bold = Font(bold=True, name='Arial', color="00800000", size=10)
        align = Alignment(horizontal="center")

        # Configuración de columnas
        headers = [
            ("A", 15, "Session ID"),
            ("B", 17, "Username"),
            ("C", 20, "Start DateTime"),
            ("D", 20, "End DateTime"),
            ("E", 15, "Time Connected (s)")
        ]
        
        for col, width, title in headers:
            self.ws.column_dimensions[col].width = width
            cell = self.ws[f"{col}1"]
            cell.font = bold
            cell.alignment = align
            cell.value = title

        self.ws.title = "History Of Their Wp"

        try:
            # Consulta para unir Users y Sessions
            query = '''
                SELECT 
                    s.id,
                    u.user_name,
                    s.start_date || ' ' || s.start_hour || ':' || s.start_minute || ':' || s.start_second AS start_datetime,
                    s.end_date || ' ' || s.end_hour || ':' || s.end_minute || ':' || s.end_second AS end_datetime,
                    s.time_connected
                FROM Sessions s
                JOIN Users u ON s.user_id = u.id
                WHERE s.end_date IS NOT NULL
                ORDER BY s.start_date DESC, s.start_hour DESC, s.start_minute DESC, s.start_second DESC
            '''
            self.cursor.execute(query)
            all_data = self.cursor.fetchall()

            # Agregar datos al Excel
            for row_idx, data in enumerate(all_data, start=2):  # Empezamos en la fila 2 por los encabezados
                self.ws[f"A{row_idx}"] = data[0]  # Session ID
                self.ws[f"B{row_idx}"] = data[1]  # Username
                self.ws[f"C{row_idx}"] = data[2]  # Start DateTime
                self.ws[f"D{row_idx}"] = data[3]  # End DateTime
                self.ws[f"E{row_idx}"] = data[4]  # Time Connected
            print("All data added to your Excel file")
            self.wb.save(self.EXCEL_FILE)

        except PermissionError:
            print(f"Please close '{self.EXCEL_FILE}' and restart the program.")
        except sqlite3.Error as e:
            print(f"Database error: {e}")
        finally:
            self.conn.close()

if __name__ == '__main__':
    converter = Converter()
    converter.db_to()
    converter.db_to_excel()